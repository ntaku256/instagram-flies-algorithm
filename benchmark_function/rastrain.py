import mpmath
import matplotlib.pyplot as plt
import matplotlib.animation as anime
import openpyxl
import random
from sklearn.cluster import KMeans
from mpl_toolkits.mplot3d import Axes3D
import numpy as np

from benchmark_function import BenchmarkFunction as BF

bf = BF()

# 評価値の計算
def Evaluate(vector):
    return bf.rosenbrock_star(vector)

# パレート分布
def Pareto(a, mode, shape): 
    print(shape)
    return [mpmath.mpf((mpmath.rand() * a) + 1) * mode for _ in range(2*1)]

# 配列の中から、値の大きさに応じた確率(大きさの比)でインデックスを選択する
def roulett(table):
    total = mpmath.mpf(0)
    for value in table:
        total += value
    rand = mpmath.mpf(mpmath.rand() * total)
    sum = mpmath.mpf(0)
    for i, value in enumerate(table):
        sum += value
        if sum > rand:
            return i

# 配列の中から、値の大きさに応じた確率(逆数の比)でインデックスを選択する
# ただし、0付近で値が極端に大きくなるのでオフセットを使用する
def roulettMin(table, offset=mpmath.mpf('1e-300')):
    total = mpmath.mpf(0)
    for value in table:
        total += mpmath.mpf(1) / (offset + value)
    rand = mpmath.mpf(mpmath.rand() * total)
    sum = mpmath.mpf(0)
    for i, value in enumerate(table):
        sum += mpmath.mpf(1) / (offset + value)
        if sum > rand:
            return i

# 粒子の位置が範囲外に出た場合は、範囲内の最大値にする
def filtIndivisual(vector, r):
    for i in range(len(vector)):
        vector[i] = max(-r, min(r, vector[i]))

# 粒子の位置を初期化
def InitIndivisual(r):
    return [mpmath.mpf(mpmath.rand() * 2 * r - r) for _ in range(2)]

def mpmath_mean(values):
    total = mpmath.mpf(0)
    count = len(values)
    if count == 0:
        return 0
    for value in values:
        total += value
    return total / count

class InstaGramFlies:
    n_iters = None
    n_clusters = None
    n_flies = None
    range = None

    # best fly in in each cluster
    best_fly_indices = None
    cluster_like_average = None
    center_dist_average = None
    center_speeds = None

    master_w = None
    master_c1 = None
    master_c2 = None
    master_c3 = None

    faddist_w = None
    faddist_c1 = None
    faddist_c2 = None
    faddist_c3 = None

    vectors = None
    centers = None
    likes = None
    labels = None
    strategies = None
    p_best_vectors = None
    p_best_score = None

    vector_strategy = None
    p_best_vector_strategy = None
    vector_cluster = None

    def __init__(self, n_iters, n_clusters, n_flies, range, master_w, master_c1, master_c2, master_c3, faddist_w, faddist_c1, faddist_c2, faddist_c3):
        self.n_iters = n_iters
        self.n_clusters = n_clusters
        self.n_flies = n_flies
        self.range = range
        self.master_w = master_w
        self.master_c1 = master_c1
        self.master_c2 = master_c2
        self.master_c3 = master_c3
        self.faddist_w = faddist_w
        self.faddist_c1 = faddist_c1
        self.faddist_c2 = faddist_c2
        self.faddist_c3 = faddist_c3
        self.InitFlies()
        self.EvaluateLikes()

    def InitFlies(self):
        self.vectors = [InitIndivisual(self.range) for _ in range(self.n_flies)]
        self.vector_strategy = ["green" for _ in range(self.n_flies)]
        self.p_best_vector_strategy = ["" for _ in range(self.n_flies)]
        self.vector_cluster = ["magenta" for _ in range(self.n_flies)]
        self.p_best_vectors = [InitIndivisual(self.range) for _ in range(self.n_flies)]
        self.p_best_score = [mpmath.inf for _ in range(self.n_flies)]
        self.likes = [mpmath.inf for _ in range(self.n_flies)]
        self.best_fly_indices = [0] * self.n_clusters
        self.InitStrategies()

    def InitStrategies(self):
        self.strategies = [[mpmath.mpf(mpmath.rand() * 100) for _ in range(3)] for _ in range(self.n_flies)]
        for i in range(self.n_flies):
            total = sum(self.strategies[i])
            self.strategies[i] = [value / total for value in self.strategies[i]]

    def ClusterColor(self):
        clusterColor = {
            0: "blue",
            1: "red",
            2: "green",
            3: "orange",
            4: "purple",
            5: "cyan",
            6: "magenta",
            7: "yellow",
            8: "black",
            9: "brown"
        }
        for i in range(self.n_flies):
            self.vector_cluster[i] = clusterColor[self.labels[i]]

    def Run(self):
        # # グラフの初期化
        # fig = plt.figure()
        # ax = Axes3D(fig)
        # ax.set_xlabel('x')
        # ax.set_ylabel('y')
        # ax.set_zlabel('f')
        # ax.view_init(elev=70, azim=0)
        # fig.add_axes(ax)
        # X = np.linspace(-self.range, self.range, 100)
        # Y = np.linspace(-self.range, self.range, 100)
        # XX, YY = np.meshgrid(X, Y)
        # d = XX.shape
        # input_array = np.vstack((XX.flatten(), YY.flatten())).T
        # Z = Evaluate(input_array)
        # Z_list = list(map(mpmath.mpf, Z))
        # d = XX.shape
        # ZZ = np.array(Z_list).reshape(d)
        # # ZZ = Z.reshape(d)
        # imgs = []


        best = []

        for i in range(self.n_iters):
            self.EvaluateLikes()
            self.Clustering()

            best_arg = self.p_best_score.index(min(self.p_best_score))
            best.append(self.p_best_score[best_arg])

            if i == 0:
                self.ClusterColor()

            # vector = np.array(self.vectors)
            # centor = np.array(self.centers)
            # # # # if (not self.centers is None) and (i < 2 or i == 4 or i == int(self.n_iters/2 -1) or i == self.n_iters -1):
            # if (not self.centers is None) and ((i - 1) < 2 or (i - 1) == 4 or (i)%100  == 0 ):
            # # if i >= 0:
            #     title = ax.text(0,0,1.8*1.61*max(Z),  't=%s' % (str(i)), size=20, zorder=1,  color='k') 
            #     scatter_vector = ax.scatter(vector.T[0], vector.T[1],Evaluate(vector),c=self.vector_cluster, s=3, alpha=0.5)
            #     # scatter_best_vector = ax.scatter(self.p_best_vectors.T[0], self.p_best_vectors.T[1],Evaluate(self.p_best_vectors),c=self.p_best_vector_strategy, s=1, alpha=0.5)
            #     scatter_func = ax.plot_surface(XX, YY, ZZ, rstride = 1, cstride = 1, cmap = plt.cm.coolwarm,alpha=0.45)
            #     if (not self.centers is None):
            #         scatter_center = ax.scatter(centor.T[0], centor.T[1],Evaluate(centor),c="black", s=10, alpha=0.8)
            #         imgs.append([title,scatter_vector,scatter_func,scatter_center])
            #     else:
            #         imgs.append([scatter_func,scatter_vector,title])

            self.UpdateFlieVector()

        self.EvaluateLikes()

        # ani = anime.ArtistAnimation(fig, imgs,interval=600)
        # ani.save("benchmark_function/GIF/insta_files/insta_files.gif",writer="imagemagick")
        # plt.show()

        best_arg = self.p_best_score.index(min(self.p_best_score))
        return self.p_best_score[best_arg], self.p_best_vectors[best_arg],best

    def EvaluateLikes(self):
        self.likes = Evaluate(self.vectors)

    def Clustering(self):
        if self.centers is None:
            model = KMeans(n_clusters=self.n_clusters)
            result = model.fit(self.vectors)
            self.centers = result.cluster_centers_
        else:
            model = KMeans(n_init=1, n_clusters=self.n_clusters, init=self.centers)
            result = model.fit(self.vectors)
        self.labels = result.labels_
        self.center_speeds = result.cluster_centers_ - self.centers
        self.centers = result.cluster_centers_

        best = [mpmath.inf] * self.n_clusters
        self.cluster_like_average = [mpmath.mpf(0) for _ in range(self.n_clusters)]
        for i in range(self.n_flies):
            self.p_best_vector_strategy[i] = "black"
            label = self.labels[i]
            if self.likes[i] < best[label]:
                best[label] = self.likes[i]
                self.best_fly_indices[label] = i
            if self.likes[i] < self.p_best_score[i]:
                self.p_best_score[i] = self.likes[i]
                self.p_best_vectors[i] = self.vectors[i]
                self.p_best_vector_strategy[i] = self.vector_strategy[i]

        for i in range(self.n_clusters):
            self.cluster_like_average[i] = mpmath_mean([mpmath.mpf(self.likes[j]) for j in range(self.n_flies) if self.labels[j] == i])

        self.center_dist_average = [mpmath.mpf(0) for _ in range(len(self.vectors[0]))]
        for i in range(self.n_clusters):
            for j in range(i+1, self.n_clusters):
                self.center_dist_average = [x + y for x, y in zip(self.center_dist_average, (self.centers[i] - self.centers[j]))]
        self.center_dist_average = [x / sum(range(1, self.n_clusters)) for x in self.center_dist_average]

    def UpdateFlieVector(self):
        for i in range(self.n_flies):
            action = roulett(self.strategies[i])
            # pioneer  or action >= 0
            # if action == 0:
            #     self.vectors[i] = self.UpdatePioneer(self.vectors[i])
            #     self.vector_strategy[i] = "green"
            # faddist
            if action == 1 or action >= 0:
                self.vectors[i] = self.UpdateFaddist(self.vectors[i])
                self.vector_strategy[i] = "blue"
            # master
            # if action == 2 or action >= 0:
            #     self.vectors[i] = self.UpdateMaster(self.vectors[i], self.labels[i],self.master_w,self.master_c1,self.master_c2,self.master_c3)
            #     self.vector_strategy[i] = "red"
            filtIndivisual(self.vectors[i],self.range)

    def UpdatePioneer(self, vector):
        center_dist_average_array = np.array(self.center_dist_average, dtype=object)
        length = Pareto(6, 1, center_dist_average_array.shape)
        length = np.array(length, dtype=object)
        rand01 = np.random.choice([-1, 1], length.shape)
        
        # Convert to mpmath.mpf
        length = np.array([mpmath.mpf(val) for val in length])
        rand01 = np.array([mpmath.mpf(val) for val in rand01])
        center_dist_average = np.array([mpmath.mpf(val) for val in self.center_dist_average])
        
        return vector + center_dist_average * length * rand01

    def UpdateFaddist(self, vector):
        cluster = roulettMin(self.cluster_like_average)
        return self.UpdateMaster(vector, cluster, self.faddist_w, self.faddist_c1, self.faddist_c2, self.faddist_c3)

    def UpdateMaster(self, vector, label, w, c1, c2, c3):
        index_table = [i for i in range(self.n_flies) if self.labels[i] == label]  
        if index_table == []:
            cluster_like_average = self.cluster_like_average
            while True: 
                cluster_like_averages = []
                for i in range(len(cluster_like_average)):
                    if i == label:
                        continue
                    cluster_like_averages.append(cluster_like_average[i])
                cluster_like_average = cluster_like_averages
                label = roulettMin(cluster_like_average)
                index_table = [i for i in range(self.n_flies) if(self.labels[i]==label)]
                if index_table != []:
                    break
        table = [mpmath.mpf(self.likes[i]) for i in index_table] # Convert likes to mpmath.mpf
        target_fly_index = index_table[roulettMin(table)]
        
        # Convert to mpmath.mpf
        center = np.array([mpmath.mpf(val) for val in self.centers[label]])
        vector = np.array([mpmath.mpf(val) for val in vector])
        target_vector = np.array([mpmath.mpf(val) for val in self.vectors[target_fly_index]])
        center_speed_vector = np.array([mpmath.mpf(val) for val in self.center_speeds[label]])

        center_vector = (center - vector) * mpmath.mpf(np.random.uniform(0, 1))
        target_vector = (target_vector - vector) * mpmath.mpf(np.random.uniform(0, 1))
        center_speed_vector = center_speed_vector * mpmath.mpf(np.random.uniform(0, 1))

        return w * vector + c1 * center_vector + c2 * target_vector + c3 * center_speed_vector

# メイン処理
if __name__ == "__main__":
    n_indivisuals = 150
    n_iters = 200
    n_clusters = 10
    r = 2.048 #5.12 or 2.048 or 100

    # 慣性
    # クラスタの中心に向かうベクトルにかける係数
    # 粒子に向かうベクトルにかける係数
    # クラスタのベクトルにかける係数

    master_w = 1
    master_c1 = 0.2 # 0.2
    master_c2 = 1.7 # 1.7
    master_c3 = 0.3 #  0.3

    faddist_w = 1 #0.4
    faddist_c1 = 0.3 #0.8
    faddist_c2 = 1.7 #0.5
    faddist_c3 = 0.2 #1.3

    # faddist_w = 1
    # faddist_c1 = 0.8
    # faddist_c2 = 0.5
    # faddist_c3 = 1.3

    insta = InstaGramFlies(n_iters, n_clusters, n_indivisuals,r,master_w,master_c1,master_c2,master_c3,faddist_w,faddist_c1,faddist_c2,faddist_c3)
    score, result,best = insta.Run()
    print(score,result)

    write_wb = openpyxl.load_workbook("results/Book_write_master.xlsx")
    write_ws = write_wb["Sheet1"]

    # シートを初期化
    for row in write_ws:
        for cell in row:
            cell.value = None

    # # c = write_ws.cell(1 , 1)
    # # c.value = "master_w"
    # # c = write_ws.cell(2 , 1)
    # # c.value = "慣性"
    # c = write_ws.cell(1 , 2)
    # c.value = "faddist_c1"
    # c = write_ws.cell(2 , 2)
    # c.value = "クラスタの中心に向かうベクトルにかける係数"
    # c = write_ws.cell(1 , 3)
    # c.value = "faddist_c2"
    # c = write_ws.cell(2 , 3)
    # c.value = "粒子に向かうベクトルにかける係数"
    # c = write_ws.cell(1 , 4)
    # c.value = "faddist_c3"
    # c = write_ws.cell(2 , 4)
    # c.value = "クラスタのベクトルにかける係数"
    # c = write_ws.cell(1 , 5)
    # c.value = "score"

    # best = []
    # faddist = []
    # count = 0
    
    # faddist_c1 = 0.1
    # for j in range(20):
    #     faddist_c2 = 0.1
    #     for k in range(20):
    #         faddist_c3 = 0.1
    #         for l in range(20):
    #             print(faddist_c1,faddist_c2,faddist_c3)
    #             # master.append([master_w,master_c1,master_c2,master_c3])
    #             insta = InstaGramFlies(n_iters, n_clusters, n_indivisuals,r,master_w,master_c1,master_c2,master_c3,faddist_w,faddist_c1,faddist_c2,faddist_c3)
    #             score, result = insta.Run()
    #             # best.append(score)
    #             result = [str(r) for r in result]
    #             score = str(score)
    #             print(score,result)

                
    #             c = write_ws.cell(count+3 , 2)
    #             c.value = faddist_c1
    #             c = write_ws.cell(count+3 , 3)
    #             c.value = faddist_c2
    #             c = write_ws.cell(count+3 , 4)
    #             c.value = faddist_c3
    #             c = write_ws.cell(count+3 , 5)
    #             c.value = faddist_c1
    #             c.value = score
    #             c = write_ws.cell(count+3 , 6)
    #             c.value = result[0]
    #             c = write_ws.cell(count+3 , 7)
    #             c.value = result[1]
    #             count += 1
            
    #             faddist_c3 += 0.1
    #         faddist_c2 += 0.1
    #     faddist_c1 += 0.1
        



    for i in range(len(best)):
        # for j in range(len(best[i])):
            c = write_ws.cell(i+1 , 1)
            c.value = str(best[i])
        
    # c = write_ws.cell(i+3 , 5)
    # c.value = best[i]
            
    write_wb.save("results/Book_write_master.xlsx")
