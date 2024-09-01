import numpy as np
import matplotlib.pyplot as plt
import matplotlib.animation as anime
import openpyxl
from sklearn.cluster import KMeans
from mpl_toolkits.mplot3d import Axes3D

from benchmark_function import BenchmarkFunction as BF

bf = BF()

# 評価
def Evaluate(vector):
    return bf.sphere(vector)

# パレート分布
def Pareto(mode,a,shape): 
    return (np.random.pareto(a,size=shape)+1)*mode

# 値が高い程、選択されやすくなるルーレット
def roulett(table):
    total = np.sum(table)
    rand = np.random.uniform(0.0,total)
    sum = 0
    for i in range(len(table)):
        sum += table[i]
        if(sum > rand):
            return i
        
# 値が低い程、選択されやすくなるルーレット
def roulettMin(table):
    total = 0
    for i in range(len(table)):
        total +=1/(1+table[i])
    rand =np.random.uniform(0.0,total)
    sum = 0
    for i in range(len(table)):
        sum +=1/table[i]
        if(sum > rand):
            return i

# 粒子の更新時に、範囲外に出たら、最大値で止める
def filtIndivisual(vector,r):
    for i in range(1*2):
        vector[i] = max(-r,min(r,vector[i]))

# 粒子の初期化
def InitIndivisual(r):
    return np.random.uniform(-r,r,(1*2))

def EvalIndivisual(vector):
    return Evaluate(vector)

class InstaGramFlies:
    n_iters = None # 試行回数
    n_clusters = None # クラスタの数
    n_flies = None # 粒子の数
    centers = None # クラスタの重心
    r = None # 範囲
    # best fly in in each cluster
    best_fly_indices = None 
    cluster_like_average = None # クラスタの平均評価値
    center_dist_average = None # クラスタの平均距離
    center_speeds = None # クラスタ全体の行動

    master_w = None
    master_c1= None
    master_c2 = None
    master_c3 = None

    faddist_w = None
    faddist_c1 = None
    faddist_c2 = None
    faddist_c3 = None

    likes = None
    labels = None
    #np.array([[pioneer rate, faddist rate, master rate],[...],...])
    strategies = None
    vectors = None  # np.array([x1,x2,x3,...]) x1: np.array
    p_best_vectors = None
    p_best_score = None
    t_g_score = [] # 試行回数に対する評価値

    def __init__(self, n_iters, n_clusters, n_flies,r,master_w,master_c1,master_c2,master_c3,faddist_w,faddist_c1,faddist_c2,faddist_c3):
        self.n_iters = n_iters
        self.n_clusters = n_clusters
        self.n_flies = n_flies
        self.r = r
        self.t_g_score = []
        self.master_w = master_w
        self.master_c1 = master_c1
        self.master_c2 = master_c2
        self.master_c3 = master_c3
        self.faddist_w = faddist_w
        self.faddist_c1 = faddist_c1
        self.faddist_c2 = faddist_c2
        self.faddist_c3 = faddist_c3
        self.InitFlies()
        self.EvaluateLikes()

    def InitFlies(self):
        self.vectors = np.array([InitIndivisual(self.r) for _ in range(self.n_flies)])
        self.p_best_vectors = np.zeros_like(self.vectors)
        self.p_best_score = np.array([float('inf') for _ in range(self.n_flies)])

        # 選択確率の決定
        self.strategies = np.zeros([self.n_flies, 3])
        for i in range(self.n_flies):
            randoms = np.random.uniform(1, 100, 3)
            for j in range(3):
                self.strategies[i][j] = randoms[j]/sum(randoms)

        self.likes = np.array([float('inf') for _ in range(self.n_flies)])
        self.best_fly_indices = np.zeros(self.n_clusters)

    def Run(self):
        # fig = plt.figure()
        # ax = Axes3D(fig)
        # ax.set_xlabel('x')
        # ax.set_ylabel('y')
        # ax.set_zlabel('f')
        # fig.add_axes(ax)
        # X = np.linspace(-self.r, self.r, 100)
        # Y = np.linspace(-self.r, self.r, 100)
        # XX, YY = np.meshgrid(X, Y)
        # d = XX.shape
        # input_array = np.vstack((XX.flatten(), YY.flatten())).T
        # Z = Evaluate(input_array)
        # ZZ = Z.reshape(d)
        # imgs = []

        for i in range(self.n_iters):
            self.EvaluateLikes()
            # if (not self.centers is None) and (i < 2 or i == 4 or i == int(self.n_iters/2 -1) or i == self.n_iters -1):
            # if (not self.centers is None) and (i < 2 or i == 4 or (i+1)%50 == 0  or i == self.n_iters - 1):
            #     title = ax.text(0,0,1.61*max(Z),  '%s' % (str(i+1)), size=20, zorder=1,  color='k') 
            #     scatter_center = ax.scatter(self.centers.T[0], self.centers.T[1],Evaluate(self.centers),c="red", s=10, alpha=0.5)
            #     # scatter_vector = ax.scatter(self.p_best_vectors.T[0], self.p_best_vectors.T[1],Evaluate(self.p_best_vectors),c="green", s=5, alpha=0.5)
            #     scatter_vector1 = ax.scatter(self.vectors.T[0], self.vectors.T[1],Evaluate(self.vectors),c="blue", s=5, alpha=0.5)
            #     scatter_func = ax.plot_surface(XX, YY, ZZ, rstride = 1, cstride = 1, cmap = plt.cm.coolwarm,alpha=0.55)
            #     imgs.append([scatter_func,scatter_center,scatter_vector1,title])

            self.Clustering()
            self.UpdateFlieVector()
            best_arg = np.argmin(self.p_best_score)
            self.t_g_score.append(self.p_best_score[best_arg])

        # ani = anime.ArtistAnimation(fig, imgs,interval=1000)
        # ani.save("benchmark_function/GIF/insta_files/insta_files.gif",writer="imagemagick")
        # plt.show()

        self.EvaluateLikes()
        # best_arg = np.argmin(self.likes)
        best_arg = np.argmin(self.p_best_score)
        return self.t_g_score,self.p_best_score[best_arg],self.vectors[best_arg]

    def EvaluateLikes(self):
        self.likes = EvalIndivisual(self.vectors)

    def Clustering(self):
        if self.centers is None:
            model = KMeans(n_clusters=self.n_clusters)
            result = model.fit(self.vectors)
            self.centers = result.cluster_centers_
        else:
            model = KMeans(n_init=1,n_clusters=self.n_clusters,init=self.centers)
            result = model.fit(self.vectors)
        self.labels = result.labels_
        self.center_speeds = result.cluster_centers_ - self.centers
        self.centers = result.cluster_centers_
    
        # best flies in each cluster
        best = np.zeros(self.n_clusters)
        self.cluster_like_average = np.zeros(self.n_clusters)
        for i in range(self.n_flies):
            label = self.labels[i]
            if (self.likes[i] < best[label]):
                best[label] = self.likes[i]
                self.best_fly_indices[label] = i
            if (self.likes[i] < self.p_best_score[i]):
                self.p_best_score[i] = self.likes[i]
                self.p_best_vectors[i] = self.vectors[i]

        # like average in each cluster
        for i in range(self.n_clusters):
            self.cluster_like_average[i] = np.mean(
                    [self.likes[j] for j in range(self.n_flies) if self.labels[j] == i])

        # average dist between each cluster
        self.center_dist_average = np.zeros_like(self.vectors[0])
        for i in range(self.n_clusters):
            for j in range(i+1,self.n_clusters):
                self.center_dist_average += (self.centers[i]-self.centers[j])
        self.center_dist_average /= sum(range(1,self.n_clusters))

    def UpdateFlieVector(self):
        for i in range(self.n_flies):
            action = roulett(self.strategies[i])
            # pioneer
            if action == 0 :
                self.vectors[i] = self.UpdatePioneer(self.vectors[i])
            # faddist
            if action == 1 :
                self.vectors[i] = self.UpdateFaddist(self.vectors[i])
            # master
            if action == 2:
                self.vectors[i] = self.UpdateMaster(self.vectors[i], self.labels[i],self.master_w,self.master_c1,self.master_c2,self.master_c3)
            filtIndivisual(self.vectors[i],self.r)

    def UpdatePioneer(self, vector):
        length = Pareto(1,6,self.center_dist_average.shape)
        rand01 = np.random.choice([-1,1],length.shape)
        return vector + self.center_dist_average*length*rand01

    def UpdateFaddist(self, vector):
        cluster = roulettMin(self.cluster_like_average)
        return self.UpdateMaster(vector,cluster,self.faddist_w,self.faddist_c1,self.faddist_c2,self.faddist_c3)

    def UpdateMaster(self, vector, label,w,c1,c2,c3):
        index_table = [i for i in range(self.n_flies) if(self.labels[i]==label)]
        table = [self.likes[i] for i in index_table]
        target_fly_index = index_table[roulettMin(table)]
        center = self.centers[label]
        center_vector = (center - vector)*np.random.uniform(0,1)
        target_vector = (self.vectors[target_fly_index] - vector)*np.random.uniform(0,1)
        center_speed_vector = self.center_speeds[label]*np.random.uniform(0,1)
        return w*vector+c1*center_vector+c2*target_vector+c3*center_speed_vector

if __name__ == "__main__":
    n_indivisuals = 150
    n_iters = 200
    n_clusters = 10
    r = 5.12 #5.12 or 2.048 or 100
    master_w =  1
    master_c1 = 1
    master_c2 = 1
    master_c3 = 1

    # faddist_w, faddist_c1, faddist_c2 ,faddist_c3 =  0.5, 0.5, 0.5, 0.5
    # faddist_c2 = 0.7
    # faddist_c3 = 0.2

    best = []
    faddist_w = 0.5
    for i in range(3):
        faddist_c1 = 0.5
        for j in range(3):
            faddist_c2 = 0.5
            for k in range(3):
                faddist_c3 = 0.5
                for l in range(3):
                    print(faddist_w,faddist_c1,faddist_c2,faddist_c3)
                    insta = InstaGramFlies(n_iters, n_clusters, n_indivisuals,r,master_w,master_c1,master_c2,master_c3,faddist_w,faddist_c1,faddist_c2,faddist_c3)
                    g_score,score, result = insta.Run()
                    best.append(g_score)
                    print(score,result)
                
                    faddist_c3 += 0.2
                faddist_c2 += 0.2
            faddist_c1 += 0.2
        faddist_w += 0.2


    write_wb = openpyxl.load_workbook("Books/Book_write.xlsx")
    write_ws = write_wb["Sheet1"]

    # シートを初期化
    for row in write_ws:
        for cell in row:
            cell.value = None

    for i in range(len(best)):
        c = write_ws.cell(1 , i+1)
        c.value = i
        for j in range(len(best[i])):
            c = write_ws.cell(j+2 , i+1)
            c.value = best[i][j]
            
    write_wb.save("Books/Book_write.xlsx")