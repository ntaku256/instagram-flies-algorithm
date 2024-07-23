import numpy as np
import matplotlib.pyplot as plt
import matplotlib.animation as anime
from sklearn.cluster import KMeans
from mpl_toolkits.mplot3d import Axes3D
import openpyxl
from benchmark_function import BenchmarkFunction as BF

bf = BF()

def Evaluate(vector):
    return bf.sphere(vector)

def Pareto(mode,a,shape):
    return (np.random.pareto(a,size=shape)+1)*mode

def roulett(table):
    total = np.sum(table)
    rand = np.random.uniform(0.0,total)
    sum = 0
    for i in range(len(table)):
        sum += table[i]
        if(sum > rand):
            return i

def filtIndivisual(vector,r):
    for i in range(1*2):
        vector[i] = max(-r,min(r,vector[i]))

def InitIndivisual(r):
    return np.random.uniform(-r,r,(1*2))

def EvalIndivisual(vector):
    return Evaluate(vector)


class PSO:
    n_iter = None
    n_swarm = None
    w = None
    c1 = None
    c2 = None
    r = None
    vectors = None
    scores = None
    g_best_score = None
    g_best_vector = None
    p_best_scores = None
    p_best_vectors = None
    j = None
    t_g_score = []

    def __init__(self,n_iter,n_swarm,w,c1,c2,r,j):
        self.n_iter = n_iter
        self.n_swarm = n_swarm
        self.w = w
        self.c1 = c1
        self.c2 = c2
        self.r = r
        self.j = j
        self.t_g_score = []
        self.InitSwarms()
        self.CalcScores()

    def InitSwarms(self):
        self.vectors = np.array([InitIndivisual(self.r) for _ in range(self.n_swarm)])
        self.speeds = np.zeros_like(self.vectors)
        self.p_best_scores = np.array([float('inf') for _ in range(self.n_swarm)])
        self.p_best_vectors = np.zeros_like(self.vectors)
        self.g_best_score = float('inf')
        self.g_best_vector= np.zeros_like(self.vectors[0])
        self.scores = np.array([float('inf') for _ in range(self.n_swarm)])
    
    def CalcScores(self):
        new_score = EvalIndivisual(self.vectors)
        for i in range(self.n_swarm):
            if new_score[i] < self.p_best_scores[i]:
                self.p_best_scores[i] = new_score[i]
                self.p_best_vectors[i] = np.copy(self.vectors[i])
            if new_score[i] < self.g_best_score:
                self.g_best_score = new_score[i]
                self.g_best_vector = np.copy(self.vectors[i])
            self.scores[i] = new_score[i]

    def UpdateVectors(self):
        for i in range(self.n_swarm):
            r1 = np.random.uniform(0,1,self.vectors[0].shape)
            r2 = np.random.uniform(0,1,self.vectors[0].shape)
            self.speeds[i] = self.w*self.speeds[i]+r1*(self.p_best_vectors[i]-self.vectors[i])+r2*(self.g_best_vector-self.vectors[i])
            self.vectors[i] = self.vectors[i] + self.speeds[i]
            filtIndivisual(self.vectors[i],self.r)
        

    def Run(self):
        fig = plt.figure()
        ax = Axes3D(fig)
        ax.set_xlabel('x')
        ax.set_ylabel('y')
        ax.set_zlabel('f')
        fig.add_axes(ax)
        X = np.linspace(-self.r, self.r, 101)
        Y = np.linspace(-self.r, self.r, 101)
        XX, YY = np.meshgrid(X, Y)
        d = XX.shape
        input_array = np.vstack((XX.flatten(), YY.flatten())).T
        Z = Evaluate(input_array)
        ZZ = Z.reshape(d)
        imgs = []

        # write_wb = openpyxl.load_workbook("Books/Book_write.xlsx")
        # write_ws = write_wb["Sheet1"]

        # シートを初期化
        # for row in write_ws:
        #     for cell in row:
        #         cell.value = None

        for i in range(self.n_iter):
            self.CalcScores()
            self.UpdateVectors()

            # c = write_ws.cell(i+1 , self.j+1)
            # c.value = self.g_best_score


            # if i < 2 or i == 4 or i == int(self.n_iter/2 - 1) or i == self.n_iter - 1:
            # if i < 2 or i == 4 or (i+1)%50 == 0  or i == self.n_iter - 1:
            #     title = ax.text(0,0,1.61*max(Z),  '%s' % (str(i+1)), size=20, zorder=1,  color='k') 
            #     scatter_vector = ax.scatter(self.vectors.T[0], self.vectors.T[1],Evaluate(self.vectors),c="green", s=10, alpha=1)
            #     scatter_func = ax.plot_surface(XX, YY, ZZ, rstride = 1, cstride = 1, cmap = plt.cm.coolwarm,alpha=0.3)
            #     imgs.append([scatter_vector,scatter_func,title])
            # print(self.g_best_score,self.g_best_vector)

            self.t_g_score.append(self.g_best_score)

            self.w = self.w - ((self.w - 0.4)/self.n_iter)
        
        # ani = anime.ArtistAnimation(fig, imgs,interval=1000)
        # ani.save("benchmark_function/GIF/pso/pso.gif",writer="imagemagick")
        # plt.show()

        # write_wb.save("Books/Book_write.xlsx")
        return self.t_g_score,self.g_best_score,self.g_best_vector

if __name__ == "__main__":
    n_indivisuals = 150
    n_iters = 1000
    c1 = 2 #0.7*2
    c2 = 2 #0.7*2
    w = 0.9
    r = 100
    best = []
    pso = PSO(n_iters,n_indivisuals,w,c1,c2,r,0)
    for i in range(10):
        pso = PSO(n_iters,n_indivisuals,w,c1,c2,r,i)
        g_score,score,result = pso.Run()
        best.append(g_score)
        print(score,result)

    write_wb = openpyxl.load_workbook("Books/Book_write.xlsx")
    write_ws = write_wb["Sheet1"]

    # シートを初期化
    for row in write_ws:
        for cell in row:
            cell.value = None

    for i in range(len(best)):
        for j in range(len(best[i])):
            c = write_ws.cell(j+1 , i+1)
            c.value = best[i][j]
            
    write_wb.save("Books/Book_write.xlsx")