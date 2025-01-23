import numpy as np
import matplotlib.pyplot as plt
import matplotlib.animation as anime
import openpyxl
import random
from sklearn.cluster import KMeans
from mpl_toolkits.mplot3d import Axes3D

from benchmark_function import BenchmarkFunction as BF

bf = BF()

# 評価値の計算
def Evaluate(vector):
    return bf.rastrigin(vector)

# パレート分布
# a: 分布の幅,小さいと大きい外れ値が出る
# mode: x軸の開始位置
# shape: 出力の型
def Pareto(a,mode,shape): 
    return (np.random.pareto(a,size=shape)+1)*mode

# 配列の中から、値の大きさに応じた確率(大きさの比)でインデックスを選択する
def roulett(table):
    rand = np.random.uniform(0.0,np.sum(table))
    sum = 0
    for i in range(len(table)):
        sum += table[i]
        if(sum > rand):
            return i

# 配列の中から、値の大きさに応じた確率(逆数の比)でインデックスを選択する
# ただし、0付近で値が極端に大きくなるのでオフセットを使用する
# オフセットの値は、求める解の精度と扱える精度を考慮して決める
def roulettMin(table, offset=1e-300):
    total = 0
    for i in range(len(table)):
        total += 1 / (offset + table[i])
    rand = random.uniform(0.0, total)
    sum = 0
    for i in range(len(table)):
        sum += 1 / (offset + table[i])
        if(sum > rand):
            return i

# 粒子の位置が範囲外に出た場合は、範囲内の最大値にする
def filtIndivisual(vector,r):
    for i in range(1*2):
        vector[i] = max(-r,min(r,vector[i]))

# 粒子の位置を初期化
def InitIndivisual(r):
    return np.random.uniform(-r,r,(1*2))

class InstaGramFlies:
    n_iters = None # 試行回数
    n_clusters = None # クラスタ数
    n_flies = None # 粒子の数
    range = None # 探索範囲

    # best fly in in each cluster
    best_fly_indices = None # クラスタ内の評価値が最も高い粒子のインデックス
    cluster_like_average = None # クラスタ内の平均評価値
    center_dist_average = None # クラスタ間の平均距離
    center_speeds = None # クラスタの速度

    master_w = None # 慣性
    master_c1= None # クラスタの中心に向かうベクトルにかける係数
    master_c2 = None # 粒子に向かうベクトルにかける係数
    master_c3 = None # クラスタのベクトルにかける係数

    faddist_w = None
    faddist_c1 = None
    faddist_c2 = None
    faddist_c3 = None

    vectors = None # 粒子の位置
    centers = None # 粒子の重心
    likes = None # 評価値の計算結果
    labels = None # それぞれの粒子がどのクラスタに所属するか
    strategies = None # 行動選択確率
    p_best_vectors = None # ベクトル内での最適解を求めた時の位置
    p_best_score = None # ベクトル内での最適解を求めた時の評価値

    vector_strategy = None # 粒子がどの行動をしたのか保存する
    p_best_vector_strategy = None # ベクトル内での最適解を求めた時の行動
    vector_cluster = None # 粒子をクラスタ事に色分け

    def __init__(self, n_iters, n_clusters, n_flies,range,master_w,master_c1,master_c2,master_c3,faddist_w,faddist_c1,faddist_c2,faddist_c3):
        self.n_iters = n_iters
        self.n_clusters = n_clusters
        self.n_flies = n_flies
        self.range = range
        self.master_w = master_w
        self.master_c1 = master_c1
        self.master_c2 = master_c2
        self.master_c3 = master_c3
        self.faddist_w = faddist_w
        self.faddist_c1 = faddist_c1
        self.faddist_c2 = faddist_c2
        self.faddist_c3 = faddist_c3
        self.InitFlies()
        self.EvaluateLikes()

    def InitFlies(self):
        self.vectors = np.array([InitIndivisual(self.range) for _ in range(self.n_flies)])
        self.vector_strategy = np.array([ "green" for _ in range(self.n_flies)])
        self.p_best_vector_strategy = np.array([ "" for _ in range(self.n_flies)])
        self.vector_cluster = np.array([ "magenta" for _ in range(self.n_flies)])
        self.p_best_vectors = np.zeros_like(self.vectors)
        self.p_best_score = np.array([float('inf') for _ in range(self.n_flies)])
        self.likes = np.array([float('inf') for _ in range(self.n_flies)])
        self.best_fly_indices = np.zeros(self.n_clusters)
        self.InitStrategies()

    # 選択行動確率の初期化
    def InitStrategies(self):
        self.strategies = np.zeros([self.n_flies, 3])
        for i in range(self.n_flies):
            randoms = np.random.uniform(1, 100, 3)
            for j in range(3):
                self.strategies[i][j] = randoms[j]/sum(randoms)
    
    def ClusterColor(self):
        clusterColor = {
            0: "blue",
            1: "red",
            2: "green",
            3: "orange",
            4: "purple",
            5: "cyan",
            6: "magenta",
            7: "yellow",
            8: "black",
            9: "brown"
        }
        for i in range(self.n_flies):
            self.vector_cluster[i] = clusterColor[self.labels[i]]

    def Run(self):
        # グラフの初期化
        fig = plt.figure()
        ax = Axes3D(fig)
        ax.set_xlabel('x')
        ax.set_ylabel('y')
        ax.set_zlabel('f')
        ax.view_init(elev=30, azim=-60)
        fig.add_axes(ax)
        X = np.linspace(-self.range, self.range, 100)
        Y = np.linspace(-self.range, self.range, 100)
        XX, YY = np.meshgrid(X, Y)
        d = XX.shape
        input_array = np.vstack((XX.flatten(), YY.flatten())).T
        Z = Evaluate(input_array)
        ZZ = Z.reshape(d)
        imgs = []

        t_best_score = []

        for i in range(self.n_iters):
            self.EvaluateLikes()
            self.Clustering()

            best_arg = np.argmin(self.p_best_score)
            t_best_score.append(self.p_best_score[best_arg])
            
            self.ClusterColor()

            # # if (not self.centers is None) and (i < 2 or i == 4 or i == int(self.n_iters/2 -1) or i == self.n_iters -1):
            if (not self.centers is None) and ((i - 1) < 2 or (i - 1) == 4 or (i)%50  == 0 ):
            # if i >= 0:
                title = ax.text(0,0,1.0*1.61*max(Z),  't=%s' % (str(i)), size=20, zorder=1,  color='k') 
                scatter_vector = ax.scatter(self.vectors.T[0], self.vectors.T[1],Evaluate(self.vectors),c=self.vector_cluster, s=3, alpha=0.5)
                # scatter_best_vector = ax.scatter(self.p_best_vectors.T[0], self.p_best_vectors.T[1],Evaluate(self.p_best_vectors),c=self.p_best_vector_strategy, s=1, alpha=0.5)
                scatter_func = ax.plot_surface(XX, YY, ZZ, rstride = 1, cstride = 1, cmap = plt.cm.coolwarm,alpha=0.55)
                if (not self.centers is None):
                    scatter_center = ax.scatter(self.centers.T[0], self.centers.T[1],Evaluate(self.centers),c="black", s=10, alpha=0.8)
                    imgs.append([title,scatter_vector,scatter_func,scatter_center])
                else:
                    imgs.append([scatter_func,scatter_vector,title])

            self.UpdateFlieVector()

        best_arg = np.argmin(self.p_best_score)
        t_best_score.append(self.p_best_score[best_arg])

        ani = anime.ArtistAnimation(fig, imgs,interval=600)
        ani.save("benchmark_function/GIF/insta_files/insta_files.gif",writer="imagemagick")
        plt.show()

        self.EvaluateLikes()

        # best_arg = np.argmin(self.likes)
        # return self.likes[best_arg],self.vectors[best_arg]
        best_arg = np.argmin(self.p_best_score)
        return self.p_best_score[best_arg],self.p_best_vectors[best_arg],t_best_score

    def EvaluateLikes(self):
        self.likes = Evaluate(self.vectors)

    def Clustering(self):
        if self.centers is None:
            model = KMeans(n_clusters=self.n_clusters) #クラスタリング(グループ分け)
            result = model.fit(self.vectors) #平均を求める
            self.centers = result.cluster_centers_
        else:
            model = KMeans(n_init=1,n_clusters=self.n_clusters,init=self.centers)
            result = model.fit(self.vectors)
        self.labels = result.labels_
        self.center_speeds = result.cluster_centers_ - self.centers
        self.centers = result.cluster_centers_
    
        # best flies in each cluster
        best = np.zeros(self.n_clusters)
        self.cluster_like_average = np.zeros(self.n_clusters)
        for i in range(self.n_flies):
            self.p_best_vector_strategy[i] = "black"
            label = self.labels[i]
            if (self.likes[i] < best[label]):
                best[label] = self.likes[i]
                self.best_fly_indices[label] = i
            if (self.likes[i] < self.p_best_score[i]):
                self.p_best_score[i] = self.likes[i]
                self.p_best_vectors[i] = self.vectors[i]
                self.p_best_vector_strategy[i] = self.vector_strategy[i]

        # like average in each cluster
        for i in range(self.n_clusters):
            self.cluster_like_average[i] = np.mean(
                    [self.likes[j] for j in range(self.n_flies) if self.labels[j] == i])
            self.cluster_like_average = np.nan_to_num(self.cluster_like_average, nan=0.0)

        # average dist between each cluster
        self.center_dist_average = np.zeros_like(self.vectors[0])
        for i in range(self.n_clusters):
            for j in range(i+1,self.n_clusters):
                self.center_dist_average += (self.centers[i]-self.centers[j])
        self.center_dist_average /= sum(range(1,self.n_clusters))

    def UpdateFlieVector(self):
        for i in range(self.n_flies):
            action = roulett(self.strategies[i])
            # pioneer  or action >= 0
            # if action == 0:
            #     self.vectors[i] = self.UpdatePioneer(self.vectors[i])
            #     self.vector_strategy[i] = "green"
            # faddist
            if action == 1 or action >= 0:
                self.vectors[i] = self.UpdateFaddist(self.vectors[i])
                self.vector_strategy[i] = "blue"
            # master
            # if action == 2 :
            #     self.vectors[i] = self.UpdateMaster(self.vectors[i], self.labels[i],self.master_w,self.master_c1,self.master_c2,self.master_c3)
            #     self.vector_strategy[i] = "red"
            filtIndivisual(self.vectors[i],self.range)

    def UpdatePioneer(self, vector):
        length = Pareto(6,1,self.center_dist_average.shape)
        rand01 = np.random.choice([-1,1],length.shape)
        return vector + self.center_dist_average*length*rand01

    def UpdateFaddist(self, vector):
        cluster = roulettMin(self.cluster_like_average)
        return self.UpdateMaster(vector,cluster,self.faddist_w,self.faddist_c1,self.faddist_c2,self.faddist_c3)

    def UpdateMaster(self, vector, label,w,c1,c2,c3):
        index_table = [i for i in range(self.n_flies) if(self.labels[i]==label)]
        # if index_table == []:
        #     cluster_like_average = self.cluster_like_average
        #     while True: 
        #         cluster_like_averages = []
        #         for i in range(len(cluster_like_average)):
        #             if i == label:
        #                 continue
        #             cluster_like_averages.append(cluster_like_average[i])
        #         cluster_like_average = cluster_like_averages
        #         label = roulettMin(cluster_like_average)
        #         index_table = [i for i in range(self.n_flies) if(self.labels[i]==label)]
        #         if index_table != []:
        #             break
        table = [self.likes[i] for i in index_table]
        target_fly_index = index_table[roulettMin(table)]
        center = self.centers[label]
        center_vector = (center - vector)*np.random.uniform(0,1)
        target_vector = (self.vectors[target_fly_index] - vector)*np.random.uniform(0,1)
        center_speed_vector = self.center_speeds[label]*np.random.uniform(0,1)
        return w*vector+c1*center_vector+c2*target_vector+c3*center_speed_vector

if __name__ == "__main__":
    n_indivisuals = 150
    n_iters = 500
    n_clusters = 10
    r = 5.12  #5.12 or 2.048 or 100

    # 慣性
    # クラスタの中心に向かうベクトルにかける係数
    # 粒子に向かうベクトルにかける係数
    # クラスタのベクトルにかける係数

    master_w =  1
    master_c1 = 0.2 #0.2
    master_c2 = 1.7 #1.7
    master_c3 = 0.3 #0.3

    faddist_w = 1 #0.4
    faddist_c1 = 0.3  #0.3 
    faddist_c2 =  1.3 #1.3
    faddist_c3 = 0.2 #0.2 

    # faddist_w = 1 #0.4
    # faddist_c1 = 1 #0.8
    # faddist_c2 = 1 #0.5
    # faddist_c3 = 1 #1.3

    # faddist_w = 1
    # faddist_c1 = 0.8
    # faddist_c2 = 0.5
    # faddist_c3 = 1.3


    insta = InstaGramFlies(n_iters, n_clusters, n_indivisuals,r,master_w,master_c1,master_c2,master_c3,faddist_w,faddist_c1,faddist_c2,faddist_c3)
    score, result,t_best_score = insta.Run()
    print(score,result)

    write_wb = openpyxl.load_workbook("results/Book_write_1.xlsx")
    write_ws = write_wb["Sheet1"]

    # シートを初期化
    for row in write_ws:
        for cell in row:
            cell.value = None

    # # c = write_ws.cell(1 , 1)
    # # c.value = "faddist_w"
    # # c = write_ws.cell(2 , 1)
    # # c.value = "慣性"
    # c = write_ws.cell(1 , 2)
    # c.value = "faddist_c1"
    # c = write_ws.cell(2 , 2)
    # c.value = "クラスタの中心に向かうベクトルにかける係数"
    # c = write_ws.cell(1 , 3)
    # c.value = "faddist_c2"
    # c = write_ws.cell(2 , 3)
    # c.value = "粒子に向かうベクトルにかける係数"
    # c = write_ws.cell(1 , 4)
    # c.value = "faddist_c3"
    # c = write_ws.cell(2 , 4)
    # c.value = "クラスタのベクトルにかける係数"
    # c = write_ws.cell(1 , 5)
    # c.value = "score"

    # best = []
    # faddist = []
    # count = 0
    
    # faddist_c1 = 0.1
    # for j in range(20):
    #     faddist_c2 = 0.1
    #     for k in range(20):
    #         faddist_c3 = 0.1
    #         for l in range(20):
    #             print(faddist_c1,faddist_c2,faddist_c3)
    #             # faddist.append([faddist_w,faddist_c1,faddist_c2,faddist_c3])
    #             insta = InstaGramFlies(n_iters, n_clusters, n_indivisuals,r,master_w,master_c1,master_c2,master_c3,faddist_w,faddist_c1,faddist_c2,faddist_c3)
    #             score, result = insta.Run()
    #             # best.append(score)
    #             print(score,result)

                
    #             c = write_ws.cell(count+3 , 2)
    #             c.value = faddist_c1
    #             c = write_ws.cell(count+3 , 3)
    #             c.value = faddist_c2
    #             c = write_ws.cell(count+3 , 4)
    #             c.value = faddist_c3
    #             c = write_ws.cell(count+3 , 5)
    #             c.value = faddist_c1
    #             c.value = score
    #             c = write_ws.cell(count+3 , 6)
    #             c.value = result[0]
    #             c = write_ws.cell(count+3 , 7)
    #             c.value = result[1]
    #             count += 1
            
    #             faddist_c3 += 0.1
    #         faddist_c2 += 0.1
    #     faddist_c1 += 0.1
        



    
    for j in range(len(t_best_score)):
        c = write_ws.cell(j+1 , 1)
        c.value = t_best_score[j]
            
            
    write_wb.save("results/Book_write_1.xlsx")
